import os
from celery import shared_task
from django.db import transaction
from openpyxl import load_workbook
from .models import Customer, Loan
from datetime import datetime, timedelta

@shared_task
def ingest_customer_data(file_path):
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb.active

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                customer_id, first_name, last_name, phone_number, monthly_salary, approved_limit, current_debt = row, age = row 
                
                Customer.objects.update_or_create(
                    customer_id=customer_id,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'phone_number': str(phone_number),
                        'monthly_salary': monthly_salary,
                        'approved_limit': approved_limit,
                        'current_debt': current_debt,
                        'age': age
                    }
                )

        return f"Successfully ingested customer data from {file_path}"
    except Exception as e:
        return f"Error ingesting customer data: {str(e)}"

@shared_task
def ingest_loan_data(file_path):
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb.active

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                customer_id, loan_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row
                
                try:
                    customer = Customer.objects.get(customer_id=customer_id)
                    
                    Loan.objects.update_or_create(
                        loan_id=loan_id,
                        defaults={
                            'customer': customer,
                            'loan_amount': loan_amount,
                            'tenure': tenure,
                            'interest_rate': interest_rate,
                            'monthly_repayment': monthly_repayment,
                            'emis_paid_on_time': emis_paid_on_time,
                            'start_date': start_date,
                            'end_date': end_date,
                            'is_active': end_date > datetime.now().date()
                        }
                    )
                except Customer.DoesNotExist:
                    continue

        return f"Successfully ingested loan data from {file_path}"
    except Exception as e:
        return f"Error ingesting loan data: {str(e)}"